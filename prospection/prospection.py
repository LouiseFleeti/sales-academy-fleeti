#!/usr/bin/env python3
"""
Fleeti France - Prospection B2B automatisee
SIRENE (INSEE) -> Pappers -> Scoring -> Export Excel
"""
import argparse, json, os, sys, time
from datetime import datetime, timedelta

try:
    import requests
except ImportError:
    print("ERREUR: pip install requests --break-system-packages"); sys.exit(1)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERREUR: pip install openpyxl --break-system-packages"); sys.exit(1)

SIRENE_BASE = "https://api.insee.fr/api-sirene/3.11"
PAPPERS_BASE = "https://api.pappers.fr/v2"

TRANCHES = {
    "NN":"Non employeur","00":"0","01":"1-2","02":"3-5","03":"6-9",
    "11":"10-19","12":"20-49","21":"50-99","22":"100-199",
    "31":"200-249","32":"250-499","41":"500-999","42":"1000-1999",
    "51":"2000-4999","52":"5000-9999","53":"10000+",
}
TRANCHE_MED = {
    "NN":0,"00":0,"01":1,"02":4,"03":7,"11":15,"12":35,"21":75,
    "22":150,"31":225,"32":375,"41":750,"42":1500,"51":3500,"52":7500,"53":15000,
}

# ── ETAPE 1: SIRENE ──

def query_sirene(token, codes_ape, departements=None, eff_min="11", eff_max="32",
                 date_apres=None, limit=100):
    headers = {"X-INSEE-Api-Key-Integration": token, "Accept": "application/json"}
    ape_f = "(" + " OR ".join(f'activitePrincipaleUniteLegale:"{c}"' for c in codes_ape) + ")"
    dep_f = ""
    if departements:
        dep_f = " AND (" + " OR ".join(f'codePostalEtablissement:{d}*' for d in departements) + ")"
    eff_f = f" AND trancheEffectifsEtablissement:[{eff_min} TO {eff_max}]"
    q = ape_f + dep_f + eff_f + ' AND etatAdministratifUniteLegale:"A" AND statutDiffusionUniteLegale:"O"'
    if date_apres:
        q += f" AND dateCreationEtablissement:[{date_apres} TO *]"

    print(f"\n{'='*60}\nETAPE 1 - Requete SIRENE\n{'='*60}")
    print(f"APE: {', '.join(codes_ape)}")
    if departements: print(f"Departements: {', '.join(departements)}")
    print(f"Effectifs: {eff_min}-{eff_max} | Limite: {limit}")
    print(f"Requete: q={q[:200]}...")

    results = []
    debut = 0
    while len(results) < limit:
        try:
            resp = requests.get(f"{SIRENE_BASE}/siret",
                headers=headers,
                params={"q": q, "nombre": min(100, limit - len(results)), "debut": debut},
                timeout=30)
        except Exception as e:
            print(f"Erreur connexion: {e}"); break

        if resp.status_code == 401:
            print("Token SIRENE invalide/expire. Regenerez-le via OAuth2."); sys.exit(1)
        elif resp.status_code == 404:
            if not results: print("Aucun resultat pour ces criteres.")
            break
        elif resp.status_code == 429:
            print("Rate limit SIRENE, pause 60s..."); time.sleep(60); continue
        elif resp.status_code != 200:
            print(f"Erreur HTTP {resp.status_code}: {resp.text[:200]}"); break

        data = resp.json()
        total = data.get("header", {}).get("total", 0)
        etabs = data.get("etablissements", [])
        if not etabs: break

        for e in etabs:
            if len(results) >= limit: break
            p = parse_etab(e)
            if p: results.append(p)

        print(f"  -> {len(results)}/{min(total, limit)} etablissements")
        debut += 100
        if debut >= total: break
        time.sleep(2.1)

    print(f"OK: {len(results)} etablissements recuperes depuis SIRENE")
    return results


def parse_etab(etab):
    ul = etab.get("uniteLegale", {})
    adr = etab.get("adresseEtablissement", {})
    per = (ul.get("periodesUniteLegale") or [{}])[0]
    denom = (ul.get("denominationUniteLegale")
             or f"{ul.get('prenomUsuelUniteLegale','')} {ul.get('nomUniteLegale','')}".strip()
             or per.get("denominationUsuelleUniteLegale","") or "Inconnu")
    cp = adr.get("codePostalEtablissement","") or ""
    commune = adr.get("libelleCommuneEtablissement","") or ""
    num = adr.get("numeroVoieEtablissement","") or ""
    tv = adr.get("typeVoieEtablissement","") or ""
    lv = adr.get("libelleVoieEtablissement","") or ""
    adresse = f"{num} {tv} {lv}, {cp} {commune}".strip().strip(",").strip()
    tr = etab.get("trancheEffectifsEtablissement","NN") or "NN"
    return {
        "siren": etab.get("siren",""), "siret": etab.get("siret",""),
        "denomination": denom, "code_ape": per.get("activitePrincipaleUniteLegale",""),
        "adresse": adresse, "code_postal": cp,
        "departement": cp[:2] if cp else "", "commune": commune,
        "tranche_code": tr, "tranche_label": TRANCHES.get(tr, tr),
        "effectif_estime": TRANCHE_MED.get(tr, 0),
        "date_creation": etab.get("dateCreationEtablissement",""),
        "cat_juridique": ul.get("categorieJuridiqueUniteLegale",""),
        "dirigeant_nom":"", "dirigeant_prenom":"", "dirigeant_fonction":"",
        "dirigeant_date_poste":"", "ca":"", "effectif_pappers":"",
        "procedure_collective": False, "site_web":"", "email_estime":"",
        "score": 0, "signaux": [],
    }

# ── ETAPE 2: PAPPERS ──

def normalize(t):
    if not t: return ""
    t = t.lower().strip()
    for o, n in {"e\u0301":"e","e\u0300":"e","e\u0302":"e","e\u0308":"e",
                 "a\u0300":"a","a\u0302":"a","u\u0300":"u","u\u0302":"u",
                 "o\u0302":"o","c\u0327":"c","i\u0302":"i","i\u0308":"i",
                 "\u00e9":"e","\u00e8":"e","\u00ea":"e","\u00eb":"e",
                 "\u00e0":"a","\u00e2":"a","\u00f9":"u","\u00fb":"u",
                 "\u00f4":"o","\u00e7":"c","\u00ee":"i","\u00ef":"i",
                 " ":"","-":"","'":""}.items():
        t = t.replace(o, n)
    return t


RNE_BASE = "https://recherche-entreprises.api.gouv.fr"

def enrich_rne(prospects, top_n=None):
    batch = prospects[:top_n] if top_n else prospects
    print(f"\n{'='*60}\nETAPE 2b - Enrichissement RNE/INPI (gratuit, {len(batch)} entreprises)\n{'='*60}")
    ok = 0
    for i, p in enumerate(batch):
        if not p["siren"]: continue
        try:
            r = requests.get(f"{RNE_BASE}/search",
                params={"q": p["siren"], "per_page": 1}, timeout=10)
        except: continue
        if r.status_code == 200:
            results = r.json().get("results", [])
            if results:
                dirs = results[0].get("dirigeants", [])
                for d in dirs:
                    if d.get("type_dirigeant") == "personne physique":
                        p["dirigeant_nom"] = d.get("nom", "") or ""
                        p["dirigeant_prenom"] = (d.get("prenoms") or "").split()[0] if d.get("prenoms") else ""
                        p["dirigeant_fonction"] = d.get("qualite", "") or ""
                        ok += 1
                        break
        elif r.status_code == 429:
            time.sleep(5); continue
        if (i+1) % 10 == 0: print(f"  -> {i+1}/{len(batch)} ({ok} enrichis)")
        time.sleep(0.5)
    print(f"OK: {ok} dirigeants recuperes via RNE (INPI)")
    return prospects


def enrich_pappers(prospects, token, top_n=None):
    if not token:
        print("\nPas de token Pappers - skip enrichissement.")
        return prospects
    batch = prospects[:top_n] if top_n else prospects
    print(f"\n{'='*60}\nETAPE 2 - Enrichissement Pappers ({len(batch)} entreprises)\n{'='*60}")
    ok = 0
    for i, p in enumerate(batch):
        if not p["siren"]: continue
        try:
            r = requests.get(f"{PAPPERS_BASE}/entreprise",
                params={"siren": p["siren"], "api_token": token}, timeout=15)
        except: continue
        if r.status_code == 200:
            d = r.json()
            reps = d.get("representants", [])
            if reps:
                p["dirigeant_nom"] = reps[0].get("nom","") or ""
                p["dirigeant_prenom"] = reps[0].get("prenom","") or ""
                p["dirigeant_fonction"] = reps[0].get("qualite","") or ""
                p["dirigeant_date_poste"] = reps[0].get("date_prise_de_poste","") or ""
            p["ca"] = d.get("chiffre_affaires","") or ""
            p["effectif_pappers"] = d.get("effectifs","") or ""
            p["site_web"] = d.get("site_web","") or ""
            p["procedure_collective"] = bool(d.get("procedure_collective"))
            ok += 1
        elif r.status_code in (401, 402):
            msg = "token invalide" if r.status_code==401 else "credits epuises"
            print(f"Pappers {msg} apres {ok} enrichissements."); break
        elif r.status_code == 429:
            time.sleep(5); continue
        if (i+1) % 10 == 0: print(f"  -> {i+1}/{len(batch)} ({ok} enrichis)")
        time.sleep(1)
    print(f"OK: {ok} enrichis via Pappers")

    # Estimation emails
    ec = 0
    for p in prospects:
        pr, nm, sw = p["dirigeant_prenom"], p["dirigeant_nom"], p["site_web"]
        if pr and nm and sw:
            dom = sw.replace("https://","").replace("http://","").replace("www.","").strip("/")
            if dom:
                pn, nn = normalize(pr), normalize(nm)
                if pn and nn:
                    p["email_estime"] = f"{pn}.{nn}@{dom}"
                    ec += 1
    print(f"Emails estimes: {ec}/{len(prospects)} (a verifier avant envoi)")
    return prospects

# ── ETAPE 3: SCORING ──

def score_prospects(prospects):
    print(f"\n{'='*60}\nETAPE 3 - Scoring\n{'='*60}")
    now = datetime.now()
    m3, m6 = now - timedelta(days=90), now - timedelta(days=180)
    for p in prospects:
        s, sig = 0, []
        dc = p.get("date_creation","")
        if dc:
            try:
                if datetime.strptime(dc, "%Y-%m-%d") >= m3: s+=3; sig.append("Creation <3 mois")
            except: pass
        dp = p.get("dirigeant_date_poste","")
        if dp:
            try:
                if datetime.strptime(dp, "%Y-%m-%d") >= m6: s+=3; sig.append("Dirigeant recent")
            except: pass
        eff = p.get("effectif_estime", 0)
        if 20 <= eff <= 200: s+=2; sig.append("Sweet spot 20-200")
        elif 10 <= eff < 20: s+=1; sig.append("PME 10-19")
        cj = str(p.get("cat_juridique",""))
        if cj in ("5710","5720","5499","5498"): s+=1; sig.append("SAS/SARL")
        if 0 < eff < 3: s-=2; sig.append("Tres petit")
        if p.get("procedure_collective"): s-=3; sig.append("Proc. collective")
        p["score"], p["signaux"] = s, sig
    prospects.sort(key=lambda x: x["score"], reverse=True)
    ch = sum(1 for p in prospects if p["score"]>=5)
    ti = sum(1 for p in prospects if 3<=p["score"]<5)
    fr = sum(1 for p in prospects if p["score"]<3)
    print(f"  Chauds (>=5): {ch}\n  Tiedes (3-4): {ti}\n  Froids (<3): {fr}")
    return prospects

# ── ETAPE 4: EXPORT EXCEL ──

def persona(eff):
    if eff < 10: return "CEO / Gerant"
    elif eff < 50: return "CEO / Resp. exploitation"
    elif eff < 200: return "COO / Fleet Manager"
    elif eff < 500: return "Fleet Manager / Resp. parc"
    return "Fleet Manager / Dir. Achats"

FONCTIONS_DECIDEURS = [
    "président", "directeur général", "gérant", "directeur general",
    "pdg", "p-dg", "ceo", "dg ", "vp ", "vice-président", "vice président",
    "managing", "associé gérant", "co-gérant", "cogérant",
    "personne ayant le pouvoir"
]

def is_decideur(fonction):
    if not fonction: return False
    f = fonction.lower()
    return any(kw in f for kw in FONCTIONS_DECIDEURS)

def export_excel(prospects, path):
    print(f"\n{'='*60}\nETAPE 4 - Export Excel\n{'='*60}")
    # Filtrer : uniquement prospects avec dirigeant décideur
    filtered = [p for p in prospects if p.get("dirigeant_nom") and is_decideur(p.get("dirigeant_fonction",""))]
    print(f"  Filtre décideurs : {len(filtered)}/{len(prospects)} prospects retenus")
    prospects = filtered if filtered else prospects  # fallback si aucun match
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects Fleeti"
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1A2A3A", end_color="1A2A3A", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hot = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    warm = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    cold = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    brd = Border(left=Side("thin","D0D0D0"), right=Side("thin","D0D0D0"),
                 top=Side("thin","D0D0D0"), bottom=Side("thin","D0D0D0"))
    hdrs = ["Prénom","Nom","Fonction","Email estimé","Entreprise","Ville"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, brd
    for ri, p in enumerate(prospects, 2):
        sc = p["score"]
        fill = hot if sc>=5 else (warm if sc>=3 else cold)
        row = [
            p["dirigeant_prenom"],
            p["dirigeant_nom"],
            p["dirigeant_fonction"],
            p["email_estime"],
            p["denomination"],
            p["commune"],
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(ri, c, v)
            cell.fill, cell.border = fill, brd
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = [20, 25, 30, 40, 40, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(prospects)+1}"

    ws2 = wb.create_sheet("Resume")
    ws2["A1"] = "Resume prospection Fleeti"
    ws2["A1"].font = Font(bold=True, size=14)
    stats = [
        ("Date extraction", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total prospects", len(prospects)),
        ("Chauds (>=5)", sum(1 for p in prospects if p["score"]>=5)),
        ("Tiedes (3-4)", sum(1 for p in prospects if 3<=p["score"]<5)),
        ("Froids (<3)", sum(1 for p in prospects if p["score"]<3)),
        ("Avec dirigeant", sum(1 for p in prospects if p["dirigeant_nom"])),
        ("Avec email estime", sum(1 for p in prospects if p["email_estime"])),
    ]
    for i, (l, v) in enumerate(stats, 3):
        ws2.cell(i, 1, l).font = Font(bold=True)
        ws2.cell(i, 2, v)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    print(f"OK: Exporte {path} ({len(prospects)} prospects)")

    # Export CSV Odoo
    import csv
    csv_path = path.replace(".xlsx", "_odoo.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Name", "First Name", "Job Position", "Email", "Company Name", "City"
        ])
        writer.writeheader()
        for p in prospects:
            writer.writerow({
                "Name": f"{p['dirigeant_prenom']} {p['dirigeant_nom']}".strip(),
                "First Name": p["dirigeant_prenom"],
                "Job Position": p["dirigeant_fonction"],
                "Email": p["email_estime"] or "",
                "Company Name": p["denomination"],
                "City": p["commune"],
            })
    print(f"OK: Export Odoo {csv_path} ({len(prospects)} contacts)")
    return path

# ── MAIN ──

def main():
    ap = argparse.ArgumentParser(description="Fleeti - Prospection B2B France")
    ap.add_argument("--sirene-token", required=True)
    ap.add_argument("--pappers-token", default="")
    ap.add_argument("--ape", nargs="+", required=True)
    ap.add_argument("--departements", nargs="*", default=None)
    ap.add_argument("--effectif-min", default="11")
    ap.add_argument("--effectif-max", default="32")
    ap.add_argument("--date-creation-apres", default=None)
    ap.add_argument("--limit", type=int, default=100)
    ap.add_argument("--enrichir-top", type=int, default=None)
    ap.add_argument("--skip-pappers", action="store_true")
    ap.add_argument("--output", default="/mnt/user-data/outputs/Prospects_Fleeti.xlsx")
    args = ap.parse_args()

    print("=" * 50)
    print("  FLEETI - Prospection B2B automatisee")
    print("  SIRENE -> Pappers -> Scoring -> Excel")
    print("=" * 50)

    prospects = query_sirene(args.sirene_token, args.ape, args.departements,
                             args.effectif_min, args.effectif_max,
                             args.date_creation_apres, args.limit)
    if not prospects:
        print("\nAucun prospect trouve."); sys.exit(0)

    prospects = enrich_rne(prospects, args.enrichir_top)
    if not args.skip_pappers and args.pappers_token:
        prospects = enrich_pappers(prospects, args.pappers_token, args.enrichir_top)
    elif not args.pappers_token and not args.skip_pappers:
        print("\nPas de token Pappers -> pas d'enrichissement dirigeants.")

    prospects = score_prospects(prospects)
    export_excel(prospects, args.output)

    print(f"\n{'='*50}\nTERMINE\n{'='*50}")
    print(f"Fichier: {args.output}")
    print(f"Total: {len(prospects)} prospects")
    top = prospects[:5]
    if top:
        print("\nTop 5:")
        for i, p in enumerate(top, 1):
            d = f"{p['dirigeant_prenom']} {p['dirigeant_nom']}".strip() or "N/A"
            print(f"  {i}. {p['denomination']} (score:{p['score']}) - {d}")

if __name__ == "__main__":
    main()
